import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
import csv
import re
import dns.resolver
from datetime import datetime, timezone

# RFC 5322 simplified regex — rejects obvious placeholders and malformed addresses
_EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")


def validate_email_format(email: str) -> str:
    """Return 'yes' if the email string matches a valid address format, 'no' otherwise.
    Template placeholders like [name]@example.com are flagged as invalid."""
    if not email or "[" in email or "]" in email:
        return "no"
    return "yes" if _EMAIL_RE.match(email.strip()) else "no"


def check_mx(domain: str) -> str:
    """Return 'yes' if the domain has at least one MX record, 'no' otherwise.
    DNS-only check — no SMTP connection is made."""
    try:
        answers = dns.resolver.resolve(domain, "MX", lifetime=10)
        return "yes" if answers else "no"
    except (
        dns.resolver.NXDOMAIN,
        dns.resolver.NoAnswer,
        dns.resolver.NoNameservers,
        dns.exception.Timeout,
        dns.resolver.LifetimeTimeout,
    ):
        return "no"

prospects = [
    # ── TIER 1: ICP match + active buying signal ──────────────────────────────
    {
        "Tier": "Tier 1",
        "Segment": "Restoration Contractor",
        "Company": "Jenkins Restorations",
        "Domain": "jenkinsrestorations.com",
        "Headcount": "200-500",
        "Fit Score": 5,
        "Buying Signal / Trigger": "Multi-state expansion (6 new offices in 18 months); hiring claims coordinators",
        "Target Contact Name": "Mike Jenkins",
        "Target Title": "President & Founder",
        "LinkedIn URL": "linkedin.com/in/mikejenkins-restoration",
        "Email Pattern": "mike@jenkinsrestorations.com",
        "Why Now": "Rapid multi-state expansion exposes claim-tracking gaps that a structured intelligence platform would immediately close.",
    },
    {
        "Tier": "Tier 1",
        "Segment": "Restoration Contractor",
        "Company": "Rytech Restoration",
        "Domain": "rytechservices.com",
        "Headcount": "100-200",
        "Fit Score": 5,
        "Buying Signal / Trigger": "Franchise model growing nationwide; seeking tech stack for franchisee claim visibility",
        "Target Contact Name": "Gary White",
        "Target Title": "CEO & Founder",
        "LinkedIn URL": "linkedin.com/company/rytech-restoration",
        "Email Pattern": "gwhite@rytechservices.com",
        "Why Now": "Franchisee network scaling creates urgent need for centralized claim intelligence to maintain brand consistency.",
    },
    {
        "Tier": "Tier 1",
        "Segment": "Restoration Contractor",
        "Company": "911 Restoration",
        "Domain": "911restoration.com",
        "Headcount": "200-500",
        "Fit Score": 5,
        "Buying Signal / Trigger": "Active Indeed postings for claims manager roles; storm season ramp-up",
        "Target Contact Name": "Idan Shpizear",
        "Target Title": "CEO & Co-Founder",
        "LinkedIn URL": "linkedin.com/in/idanshpizear",
        "Email Pattern": "idan@911restoration.com",
        "Why Now": "Hiring surge for claims staff signals process pain; ClaimSignal replaces headcount with intelligence.",
    },
    {
        "Tier": "Tier 1",
        "Segment": "Public Adjuster",
        "Company": "Five Star Claims Adjusting",
        "Domain": "fivestarclaimsadjusting.com",
        "Headcount": "10-50",
        "Fit Score": 5,
        "Buying Signal / Trigger": "Posted 3 adjuster roles in Q1 2026; Florida market post-hurricane surge",
        "Target Contact Name": "Victor Morgado",
        "Target Title": "Owner & Principal Adjuster",
        "LinkedIn URL": "linkedin.com/company/five-star-claims-adjusting",
        "Email Pattern": "victor@fivestarclaimsadjusting.com",
        "Why Now": "Florida's post-hurricane surge is overwhelming manual adjuster workflows — friction scoring would cut adjuster re-inspection rates.",
    },
    {
        "Tier": "Tier 1",
        "Segment": "Public Adjuster",
        "Company": "Greenspan Co / Adjusters International",
        "Domain": "thegreenspanco.com",
        "Headcount": "50-100",
        "Fit Score": 5,
        "Buying Signal / Trigger": "Actively seeking claims data software on LinkedIn; California wildfire caseload spike",
        "Target Contact Name": "Scott Greenspan",
        "Target Title": "President",
        "LinkedIn URL": "linkedin.com/company/the-greenspan-co",
        "Email Pattern": "scott@thegreenspanco.com",
        "Why Now": "California wildfire caseload spike demands audit-ready documentation — ClaimSignal's evidence pipeline fits exactly.",
    },
    {
        "Tier": "Tier 1",
        "Segment": "Public Adjuster",
        "Company": "ClaimsMate",
        "Domain": "claimsmate.com",
        "Headcount": "10-50",
        "Fit Score": 5,
        "Buying Signal / Trigger": "SaaS-forward brand; reviewing tools on G2; recently added paid blog content on claim supplementing",
        "Target Contact Name": "J.P. Moreau",
        "Target Title": "Founder & CEO",
        "LinkedIn URL": "linkedin.com/company/claimsmate",
        "Email Pattern": "jp@claimsmate.com",
        "Why Now": "Tech-forward firm already investing in SaaS tooling; ClaimSignal supplements their client-facing brand with backend intelligence.",
    },
    {
        "Tier": "Tier 1",
        "Segment": "Restoration Contractor",
        "Company": "Blackmon Mooring & BMS CAT",
        "Domain": "bmscat.com",
        "Headcount": "500+",
        "Fit Score": 4,
        "Buying Signal / Trigger": "Catastrophe response team scaling; active FEMA deployments and insurance carrier partnerships",
        "Target Contact Name": "Patrick Rafferty",
        "Target Title": "VP of Operations",
        "LinkedIn URL": "linkedin.com/company/bms-cat",
        "Email Pattern": "prafferty@bmscat.com",
        "Why Now": "FEMA deployment volumes create adjuster bottlenecks that lifecycle phase tracking and escalation architecture directly solve.",
    },
    {
        "Tier": "Tier 1",
        "Segment": "Insurance Consulting",
        "Company": "Envista Forensics",
        "Domain": "envistaforensics.com",
        "Headcount": "200-500",
        "Fit Score": 5,
        "Buying Signal / Trigger": "Opened 4 new US offices in 2025; hiring claims investigators and forensic engineers",
        "Target Contact Name": "Jeffrey Gnerer",
        "Target Title": "CEO",
        "LinkedIn URL": "linkedin.com/company/envista-forensics",
        "Email Pattern": "jgnerer@envistaforensics.com",
        "Why Now": "Rapid office expansion means inconsistent claim documentation standards — ClaimSignal creates uniformity across locations.",
    },
    {
        "Tier": "Tier 1",
        "Segment": "Insurance Consulting",
        "Company": "JS Held",
        "Domain": "jsheld.com",
        "Headcount": "500+",
        "Fit Score": 4,
        "Buying Signal / Trigger": "Series of acquisitions expanding property claims practice; seeking integrated platforms",
        "Target Contact Name": "Jonathon Held",
        "Target Title": "Co-CEO",
        "LinkedIn URL": "linkedin.com/company/js-held",
        "Email Pattern": "jheld@jsheld.com",
        "Why Now": "Acquisition-driven growth requires a unifying data layer — ClaimSignal's multi-org architecture fits their consulting federation model.",
    },
    {
        "Tier": "Tier 1",
        "Segment": "Restoration Contractor",
        "Company": "United Water Restoration",
        "Domain": "unitedwaterrestoration.com",
        "Headcount": "50-100",
        "Fit Score": 5,
        "Buying Signal / Trigger": "Expanding from Florida base into Southeast; actively posting operations roles",
        "Target Contact Name": "Anthony Kammas",
        "Target Title": "President & Founder",
        "LinkedIn URL": "linkedin.com/company/united-water-restoration-group",
        "Email Pattern": "anthony@unitedwaterrestoration.com",
        "Why Now": "Southeast expansion into high-storm markets creates immediate need for claim friction and adjuster intelligence.",
    },
    {
        "Tier": "Tier 1",
        "Segment": "Public Adjuster",
        "Company": "Adjusters International",
        "Domain": "adjustersinternational.com",
        "Headcount": "100-200",
        "Fit Score": 5,
        "Buying Signal / Trigger": "NAPIA board-level company; thought leadership content signals openness to tech; large commercial caseloads",
        "Target Contact Name": "Michael Loughlin",
        "Target Title": "President",
        "LinkedIn URL": "linkedin.com/company/adjusters-international",
        "Email Pattern": "mloughlin@adjustersinternational.com",
        "Why Now": "Commercial large-loss claims have complex supplement cycles — ClaimSignal's Scope Delta and Escalation engines add direct value.",
    },
    {
        "Tier": "Tier 1",
        "Segment": "Restoration Contractor",
        "Company": "Restoration 1",
        "Domain": "restoration1.com",
        "Headcount": "200-500",
        "Fit Score": 4,
        "Buying Signal / Trigger": "300+ franchise units nationwide; franchise owner satisfaction tied to claim cycle times",
        "Target Contact Name": "Gary Findley",
        "Target Title": "CEO",
        "LinkedIn URL": "linkedin.com/company/restoration-1",
        "Email Pattern": "gfindley@restoration1.com",
        "Why Now": "Franchise network depends on predictable claim outcomes; ClaimSignal's intelligence layer standardizes performance across all units.",
    },
    # ── TIER 2: Good ICP fit, no specific active trigger ──────────────────────
    {
        "Tier": "Tier 2",
        "Segment": "Restoration Contractor",
        "Company": "Rainbow Restoration",
        "Domain": "rainbowrestoration.com",
        "Headcount": "200-500",
        "Fit Score": 4,
        "Buying Signal / Trigger": "Franchise expansion ongoing; Neighborly brand portfolio",
        "Target Contact Name": "Jeff Dudan",
        "Target Title": "CEO",
        "LinkedIn URL": "linkedin.com/company/rainbow-restoration",
        "Email Pattern": "jdudan@rainbowrestoration.com",
        "Why Now": "Part of Neighborly's service franchise portfolio scaling nationally — claim intel helps franchisees compete on outcome quality.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Restoration Contractor",
        "Company": "Maxons Restorations",
        "Domain": "maxons.com",
        "Headcount": "50-100",
        "Fit Score": 4,
        "Buying Signal / Trigger": "NYC/Tri-state urban market; commercial high-rise specialty",
        "Target Contact Name": "Marc Glassberg",
        "Target Title": "President & Owner",
        "LinkedIn URL": "linkedin.com/company/maxons-restorations",
        "Email Pattern": "marc@maxons.com",
        "Why Now": "High-value commercial claims in NYC require airtight documentation — PII-protected audit trails are a direct differentiator.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Restoration Contractor",
        "Company": "AdvantaClean",
        "Domain": "advantaclean.com",
        "Headcount": "100-200",
        "Fit Score": 4,
        "Buying Signal / Trigger": "Home Franchise Concepts portfolio; mold/water/air specialty",
        "Target Contact Name": "Craig Mileham",
        "Target Title": "COO",
        "LinkedIn URL": "linkedin.com/company/advantaclean",
        "Email Pattern": "cmileham@advantaclean.com",
        "Why Now": "Mold and air quality claims have long supplement cycles; escalation modeling gives adjusters fewer comebacks.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Restoration Contractor",
        "Company": "Steamatic",
        "Domain": "steamatic.com",
        "Headcount": "50-200",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Fire/smoke specialty restoration with insurance-direct billing",
        "Target Contact Name": "Joe Martinelli",
        "Target Title": "President",
        "LinkedIn URL": "linkedin.com/company/steamatic",
        "Email Pattern": "joe@steamatic.com",
        "Why Now": "Fire/smoke claims are high-friction with carriers — friction scoring turns anecdotal knowledge into defensible data.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Restoration Contractor",
        "Company": "DKI Services",
        "Domain": "dkiservices.com",
        "Headcount": "200-500",
        "Fit Score": 4,
        "Buying Signal / Trigger": "Canada/US hybrid; US restoration network of 600+ crews",
        "Target Contact Name": "Dan Cassidy",
        "Target Title": "President, US Operations",
        "LinkedIn URL": "linkedin.com/company/dki-services",
        "Email Pattern": "dcassidy@dkiservices.com",
        "Why Now": "US network scale creates data blind spots across crews — ClaimSignal's org-level dashboards unify claim performance.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Public Adjuster",
        "Company": "Globe Midwest / Adjusters International",
        "Domain": "globemidwest.com",
        "Headcount": "50-100",
        "Fit Score": 4,
        "Buying Signal / Trigger": "Midwest tornado alley presence; large commercial property focus",
        "Target Contact Name": "Steven Severance",
        "Target Title": "President",
        "LinkedIn URL": "linkedin.com/company/globe-midwest-adjusters-international",
        "Email Pattern": "sseverance@globemidwest.com",
        "Why Now": "Commercial tornado/weather claims have unpredictable supplement cycles — outcome migration scoring helps forecast settlements.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Public Adjuster",
        "Company": "Levy, Von Beck, Comstock & Associates",
        "Domain": "lvbpa.com",
        "Headcount": "10-50",
        "Fit Score": 4,
        "Buying Signal / Trigger": "Pacific Northwest PA firm with regulatory expertise; complex commercial claims",
        "Target Contact Name": "Michael Von Beck",
        "Target Title": "Principal",
        "LinkedIn URL": "linkedin.com/company/levy-von-beck-comstock",
        "Email Pattern": "mvonbeck@lvbpa.com",
        "Why Now": "Regulatory complexity in PNW requires documented adjuster behavior trails — ClaimSignal's audit log is a compliance asset.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Public Adjuster",
        "Company": "National Fire Adjustment Co.",
        "Domain": "nfaco.com",
        "Headcount": "50-100",
        "Fit Score": 5,
        "Buying Signal / Trigger": "100+ year old firm modernizing; multi-office US presence; large loss specialty",
        "Target Contact Name": "David Barrack",
        "Target Title": "President",
        "LinkedIn URL": "linkedin.com/company/national-fire-adjustment-co",
        "Email Pattern": "dbarrack@nfaco.com",
        "Why Now": "Legacy firm investing in modernization — ClaimSignal positions as the intelligence layer for a tech-forward rebrand.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Public Adjuster",
        "Company": "Scott Adjusters",
        "Domain": "scottadjusters.com",
        "Headcount": "10-50",
        "Fit Score": 4,
        "Buying Signal / Trigger": "Southeast US focus; homeowner advocacy brand; property damage specialty",
        "Target Contact Name": "Michael Scott",
        "Target Title": "Owner & Principal",
        "LinkedIn URL": "linkedin.com/company/scott-adjusters",
        "Email Pattern": "michael@scottadjusters.com",
        "Why Now": "SE homeowner-focused brand benefits from PII-protected claim dashboards that strengthen client trust.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Public Adjuster",
        "Company": "Vanguard Claims Management",
        "Domain": "vcmgrp.com",
        "Headcount": "10-50",
        "Fit Score": 4,
        "Buying Signal / Trigger": "Texas/Gulf Coast PA firm; hail and hurricane specialty",
        "Target Contact Name": "Robert Anderson",
        "Target Title": "Managing Partner",
        "LinkedIn URL": "linkedin.com/company/vanguard-claims-management",
        "Email Pattern": "randerson@vcmgrp.com",
        "Why Now": "Gulf Coast hail season 2026 creating claims surge — real-time friction scoring accelerates adjuster triage.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Insurance Consulting",
        "Company": "Rimkus Consulting Group",
        "Domain": "rimkus.com",
        "Headcount": "500+",
        "Fit Score": 4,
        "Buying Signal / Trigger": "Engineering/forensic consulting with insurance claims as core offering; 20+ US offices",
        "Target Contact Name": "Robert Hendricks",
        "Target Title": "CEO",
        "LinkedIn URL": "linkedin.com/company/rimkus-consulting-group",
        "Email Pattern": "rhendricks@rimkus.com",
        "Why Now": "Engineering opinions on property claims are only as strong as the behavioral data behind them — ClaimSignal provides that layer.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Insurance Consulting",
        "Company": "Donan Engineering",
        "Domain": "donan.com",
        "Headcount": "200-500",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Forensic engineering with insurance claim causation focus; Midwest/national reach",
        "Target Contact Name": "Mike Donan",
        "Target Title": "President",
        "LinkedIn URL": "linkedin.com/company/donan-engineering",
        "Email Pattern": "mdonan@donan.com",
        "Why Now": "Causation investigations feed directly into supplement and denial decisions — lifecycle phase tracking integrates naturally.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Insurance Consulting",
        "Company": "McLarens",
        "Domain": "mclarens.com",
        "Headcount": "500+",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Global TPA/consulting with US property claims practice; carrier-adjacent but contractor-facing divisions",
        "Target Contact Name": "Tim Rayner",
        "Target Title": "CEO, Americas",
        "LinkedIn URL": "linkedin.com/company/mclarens",
        "Email Pattern": "tim.rayner@mclarens.com",
        "Why Now": "US property division serves contractors and consultants who need behavioral analytics to supplement carrier data.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Restoration Contractor",
        "Company": "Monarch Restoration",
        "Domain": "monarchrestoration.com",
        "Headcount": "10-50",
        "Fit Score": 4,
        "Buying Signal / Trigger": "Boutique high-end Colorado restoration contractor; growing referral network",
        "Target Contact Name": "Chris Hanson",
        "Target Title": "Owner",
        "LinkedIn URL": "linkedin.com/company/monarch-restoration",
        "Email Pattern": "chris@monarchrestoration.com",
        "Why Now": "High-value mountain-region claims with complex moisture and mold scopes need structured supplementing tools.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Restoration Contractor",
        "Company": "Lotus Restoration Services",
        "Domain": "lotusrestoration.com",
        "Headcount": "10-50",
        "Fit Score": 4,
        "Buying Signal / Trigger": "Pacific Northwest mold/water specialist; insurance-direct billing model",
        "Target Contact Name": "Dan Rowe",
        "Target Title": "Owner & President",
        "LinkedIn URL": "linkedin.com/company/lotus-restoration-services",
        "Email Pattern": "dan@lotusrestoration.com",
        "Why Now": "PNW moisture claims are high-supplement and high-friction — friction scoring gives owners leverage in carrier negotiations.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Restoration Contractor",
        "Company": "MidAmerica Restoration",
        "Domain": "midamericarestoration.com",
        "Headcount": "10-50",
        "Fit Score": 4,
        "Buying Signal / Trigger": "Oklahoma/Midwest tornado belt contractor; seasonal volume spikes",
        "Target Contact Name": "Brad Ellis",
        "Target Title": "Owner",
        "LinkedIn URL": "linkedin.com/company/midamerica-restoration",
        "Email Pattern": "brad@midamericarestoration.com",
        "Why Now": "Tornado season 2026 claim surge needs structured intake — ClaimSignal's evidence pipeline replaces spreadsheet chaos.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Public Adjuster",
        "Company": "Premier Claims",
        "Domain": "premierclaims.com",
        "Headcount": "10-50",
        "Fit Score": 4,
        "Buying Signal / Trigger": "Nebraska-based hail/storm PA; aggressive growth in Midwest",
        "Target Contact Name": "Tyler Huck",
        "Target Title": "Founder & CEO",
        "LinkedIn URL": "linkedin.com/in/tylerhuck",
        "Email Pattern": "tyler@premierclaims.com",
        "Why Now": "Hail-season surge requires fast claim triage — escalation architecture would cut the time from intake to settlement offer.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Public Adjuster",
        "Company": "Stellar Public Adjusting",
        "Domain": "stellarpa.com",
        "Headcount": "1-10",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Solo-to-small PA firm in Florida; high homeowner claim volume",
        "Target Contact Name": "Jason Rivera",
        "Target Title": "Principal Adjuster",
        "LinkedIn URL": "linkedin.com/company/stellar-public-adjusting",
        "Email Pattern": "jason@stellarpa.com",
        "Why Now": "Solo PA firms competing with larger firms need an intelligence edge — Founding Partner tier at $99/mo is accessible.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Insurance Consulting",
        "Company": "Young & Associates",
        "Domain": "younginc.com",
        "Headcount": "50-200",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Insurance regulatory and management consulting; property claims advisory",
        "Target Contact Name": "Robert Young",
        "Target Title": "CEO",
        "LinkedIn URL": "linkedin.com/company/young-associates-inc",
        "Email Pattern": "ryoung@younginc.com",
        "Why Now": "Regulatory consulting firms increasingly need audit-ready claim documentation to advise contractor clients defensively.",
    },
    {
        "Tier": "Tier 2",
        "Segment": "Insurance Consulting",
        "Company": "Exponent",
        "Domain": "exponent.com",
        "Headcount": "500+",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Engineering/science consulting with property damage practice; litigation-support focus",
        "Target Contact Name": "Paul Sheridan",
        "Target Title": "Managing Director, Insurance",
        "LinkedIn URL": "linkedin.com/company/exponent",
        "Email Pattern": "psheridan@exponent.com",
        "Why Now": "Litigation-support practice needs structured event timelines — ClaimSignal's intelligence events map directly to legal discovery.",
    },
    # ── TIER 3: Partial fit ────────────────────────────────────────────────────
    {
        "Tier": "Tier 3",
        "Segment": "Restoration Contractor",
        "Company": "BELFOR Property Restoration",
        "Domain": "belfor.com",
        "Headcount": "500+",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Global enterprise; US market leader; may already have internal tooling",
        "Target Contact Name": "Sheldon Yellen",
        "Target Title": "CEO",
        "LinkedIn URL": "linkedin.com/company/belfor",
        "Email Pattern": "syellen@belfor.com",
        "Why Now": "Enterprise scale means divisional adoption is realistic even without top-down mandate — target a regional VP first.",
    },
    {
        "Tier": "Tier 3",
        "Segment": "Restoration Contractor",
        "Company": "Paul Davis Restoration",
        "Domain": "pauldavis.com",
        "Headcount": "500+",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Large franchise; franchise owners are independent buyers",
        "Target Contact Name": "Jeff Drury",
        "Target Title": "President",
        "LinkedIn URL": "linkedin.com/company/paul-davis-restoration",
        "Email Pattern": "jdrury@pauldavis.com",
        "Why Now": "Individual franchise owners inside Paul Davis are independent decision-makers — target them directly not corporate.",
    },
    {
        "Tier": "Tier 3",
        "Segment": "Restoration Contractor",
        "Company": "ServiceMaster Restore",
        "Domain": "servicemaster.com",
        "Headcount": "500+",
        "Fit Score": 2,
        "Buying Signal / Trigger": "Large public company; complex buying cycle; franchise owner angle possible",
        "Target Contact Name": "Franchise Owner (varies)",
        "Target Title": "Franchise Owner",
        "LinkedIn URL": "linkedin.com/company/servicemaster",
        "Email Pattern": "[owner]@smrestore[location].com",
        "Why Now": "Individual franchise owners have budget autonomy and the same claim pain as independent contractors.",
    },
    {
        "Tier": "Tier 3",
        "Segment": "Restoration Contractor",
        "Company": "Interstate Restoration",
        "Domain": "interstaterestoration.com",
        "Headcount": "200-500",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Large-loss commercial specialty; multi-state US presence",
        "Target Contact Name": "Troy Moraes",
        "Target Title": "CEO",
        "LinkedIn URL": "linkedin.com/company/interstate-restoration",
        "Email Pattern": "tmoraes@interstaterestoration.com",
        "Why Now": "Large-loss commercial claims are high-value and high-complexity — structured lifecycle tracking is a margin-protection tool.",
    },
    {
        "Tier": "Tier 3",
        "Segment": "Restoration Contractor",
        "Company": "Servpro Industries (Franchise Owner)",
        "Domain": "servpro.com",
        "Headcount": "1-50 (per franchise)",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Individual franchisees are independent; corporate not a target",
        "Target Contact Name": "Local Franchise Owner",
        "Target Title": "Owner",
        "LinkedIn URL": "linkedin.com/company/servpro",
        "Email Pattern": "[name]@servpro[number].com",
        "Why Now": "Franchisee owners operate as independent businesses — high volume of property claims with limited analytics support.",
    },
    {
        "Tier": "Tier 3",
        "Segment": "Public Adjuster",
        "Company": "Keys Claims Consultants",
        "Domain": "keysclaimscc.com",
        "Headcount": "10-50",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Florida Keys specialty; hurricane/flood focus; boutique clientele",
        "Target Contact Name": "Chad Hett",
        "Target Title": "Principal",
        "LinkedIn URL": "linkedin.com/company/keys-claims-consultants",
        "Email Pattern": "chad@keysclaimscc.com",
        "Why Now": "Boutique Florida PA firms need differentiation post-Ian/Idalia — behavioral analytics is a client-retention story.",
    },
    {
        "Tier": "Tier 3",
        "Segment": "Public Adjuster",
        "Company": "Property Damage Appraisers (PDA)",
        "Domain": "pda.com",
        "Headcount": "50-200",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Appraisal and umpire services; adjacent to PA market",
        "Target Contact Name": "Dave Wehr",
        "Target Title": "President",
        "LinkedIn URL": "linkedin.com/company/property-damage-appraisers",
        "Email Pattern": "dwehr@pda.com",
        "Why Now": "Appraisal firms involved in dispute resolution need claim behavior data to support umpire positions.",
    },
    {
        "Tier": "Tier 3",
        "Segment": "Insurance Consulting",
        "Company": "Haag Engineering",
        "Domain": "haag.com",
        "Headcount": "200-500",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Roof/hail forensic engineering; carrier-facing but contractor clients exist",
        "Target Contact Name": "Chris Haag",
        "Target Title": "President",
        "LinkedIn URL": "linkedin.com/company/haag-engineering",
        "Email Pattern": "chaag@haag.com",
        "Why Now": "Engineering firms supporting contractor supplement claims need data trails — ClaimSignal provides the behavioral context.",
    },
    {
        "Tier": "Tier 3",
        "Segment": "Insurance Consulting",
        "Company": "Sedgwick (contractor-facing division)",
        "Domain": "sedgwick.com",
        "Headcount": "500+",
        "Fit Score": 2,
        "Buying Signal / Trigger": "TPA giant; contractor vendor management is a secondary touchpoint",
        "Target Contact Name": "Mike Arbour",
        "Target Title": "EVP, Property Solutions",
        "LinkedIn URL": "linkedin.com/company/sedgwick",
        "Email Pattern": "mike.arbour@sedgwick.com",
        "Why Now": "Vendor management teams inside Sedgwick influence contractor software decisions — a pilot conversation is low-risk.",
    },
    {
        "Tier": "Tier 3",
        "Segment": "Insurance Consulting",
        "Company": "Safelite Solutions (property claims consulting)",
        "Domain": "safelsolutions.com",
        "Headcount": "100-200",
        "Fit Score": 2,
        "Buying Signal / Trigger": "Managed repair network consulting; property adjacent",
        "Target Contact Name": "Brian O'Neill",
        "Target Title": "VP, Property",
        "LinkedIn URL": "linkedin.com/company/safelite-solutions",
        "Email Pattern": "boneill@safelsolutions.com",
        "Why Now": "Managed repair networks need contractor performance data — ClaimSignal's adjuster aggregated metrics are a natural fit.",
    },
    {
        "Tier": "Tier 3",
        "Segment": "Restoration Contractor",
        "Company": "Coastal Restoration Group",
        "Domain": "coastalrestorationgroup.com",
        "Headcount": "10-50",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Carolinas/Gulf Coast specialty; hurricane season 2026 exposure",
        "Target Contact Name": "Rick Alvarez",
        "Target Title": "Owner",
        "LinkedIn URL": "linkedin.com/company/coastal-restoration-group",
        "Email Pattern": "rick@coastalrestorationgroup.com",
        "Why Now": "Hurricane season start means claim intake is about to spike — ClaimSignal onboarding now gets ahead of the surge.",
    },
    {
        "Tier": "Tier 3",
        "Segment": "Restoration Contractor",
        "Company": "Fire & Ice Restoration",
        "Domain": "fireandiceinc.com",
        "Headcount": "10-50",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Minnesota/Great Lakes specialist; fire and ice damage focus",
        "Target Contact Name": "Tom Larson",
        "Target Title": "President & Owner",
        "LinkedIn URL": "linkedin.com/company/fire-and-ice-restoration",
        "Email Pattern": "tom@fireandiceinc.com",
        "Why Now": "Ice dam and fire claims in cold climates have high supplement frequency — friction scoring adds negotiating leverage.",
    },
    {
        "Tier": "Tier 3",
        "Segment": "Public Adjuster",
        "Company": "Advocate Claims Service",
        "Domain": "advocateclaims.com",
        "Headcount": "1-10",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Solo PA; referral-based; Texas/Oklahoma focus",
        "Target Contact Name": "Jennifer Hall",
        "Target Title": "Owner & PA",
        "LinkedIn URL": "linkedin.com/company/advocate-claims-service",
        "Email Pattern": "jennifer@advocateclaims.com",
        "Why Now": "Solo PAs compete on outcomes — ClaimSignal's Founding Partner tier gives a solo operator enterprise-grade claim intelligence.",
    },
    {
        "Tier": "Tier 3",
        "Segment": "Public Adjuster",
        "Company": "The People's Choice PA",
        "Domain": "thepeopleschoicepa.com",
        "Headcount": "1-10",
        "Fit Score": 3,
        "Buying Signal / Trigger": "Client-advocacy brand; social media active; Florida-based",
        "Target Contact Name": "Maria Santos",
        "Target Title": "Founder & PA",
        "LinkedIn URL": "linkedin.com/company/peoples-choice-pa",
        "Email Pattern": "maria@thepeopleschoicepa.com",
        "Why Now": "Social-media-forward PA brand can showcase ClaimSignal dashboards as a client transparency differentiator.",
    },
]

# ── Column order for base output ──────────────────────────────────────────────
BASE_COLUMNS = [
    "Tier", "Segment", "Company", "Domain", "Headcount", "Fit Score",
    "Buying Signal / Trigger", "Target Contact Name", "Target Title",
    "LinkedIn URL", "Email Pattern", "Why Now",
]

COLUMNS = BASE_COLUMNS + ["Email Format Valid", "Domain MX Valid", "Verified At"]

# Sort: Tier 1 → Tier 2 → Tier 3, then by Fit Score desc
tier_order = {"Tier 1": 0, "Tier 2": 1, "Tier 3": 2}
prospects.sort(key=lambda r: (tier_order[r["Tier"]], -r["Fit Score"]))

# ── Tier color map ─────────────────────────────────────────────────────────────
TIER_FILLS = {
    "Tier 1": PatternFill(start_color="1E4D2B", end_color="1E4D2B", fill_type="solid"),  # dark green
    "Tier 2": PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid"),  # dark blue
    "Tier 3": PatternFill(start_color="3D2B1A", end_color="3D2B1A", fill_type="solid"),  # dark amber
}
TIER_FONT_COLORS = {
    "Tier 1": "C6EFCE",  # light green
    "Tier 2": "BDD7EE",  # light blue
    "Tier 3": "FFDAB9",  # light peach
}

# ── Write base CSV (without verification columns) ─────────────────────────────
CSV_PATH = "claimsignal_prospects.csv"
with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=BASE_COLUMNS)
    writer.writeheader()
    for p in prospects:
        writer.writerow({col: p[col] for col in BASE_COLUMNS})

print(f"✓ Base CSV written — {len(prospects)} rows")

# ── Verification pass: reads CSV, adds email + MX checks, writes back ─────────
print("\nRunning email format validation and MX record checks (DNS-only, no SMTP)...")
verified_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

with open(CSV_PATH, newline="", encoding="utf-8") as f:
    rows = list(csv.DictReader(f))

seen_domains: dict[str, str] = {}
for row in rows:
    row["Email Format Valid"] = validate_email_format(row.get("Email Pattern", ""))

    domain = row["Domain"]
    if domain not in seen_domains:
        mx_result = check_mx(domain)
        seen_domains[domain] = mx_result
        fmt_status = "✓" if row["Email Format Valid"] == "yes" else "✗"
        mx_status = "✓" if mx_result == "yes" else "✗"
        print(
            f"  email {fmt_status}  mx {mx_status}  "
            f"{row['Email Pattern']:45s}  [{domain}]"
        )
    row["Domain MX Valid"] = seen_domains[domain]
    row["Verified At"] = verified_at

with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=COLUMNS)
    writer.writeheader()
    writer.writerows(rows)

email_valid_count = sum(1 for r in rows if r["Email Format Valid"] == "yes")
email_invalid_count = len(rows) - email_valid_count
mx_valid_count = sum(1 for r in rows if r["Domain MX Valid"] == "yes")
mx_invalid_count = len(rows) - mx_valid_count

print(f"\nEmail format: {email_valid_count} valid, {email_invalid_count} flagged")
print(f"Domain MX:    {mx_valid_count} valid, {mx_invalid_count} flagged\n")
print(f"✓ Verified CSV saved — {len(rows)} rows")

# Merge verified columns back into the in-memory prospects for XLSX
verified_by_domain_email: dict[tuple, dict] = {
    (r["Domain"], r["Email Pattern"]): r for r in rows
}
for p in prospects:
    key = (p["Domain"], p["Email Pattern"])
    if key in verified_by_domain_email:
        vrow = verified_by_domain_email[key]
        p["Email Format Valid"] = vrow["Email Format Valid"]
        p["Domain MX Valid"] = vrow["Domain MX Valid"]
        p["Verified At"] = vrow["Verified At"]

# ── Build XLSX ────────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Prospects"

# Header style
header_fill = PatternFill(start_color="0D1F3C", end_color="0D1F3C", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="444444")
header_border = Border(bottom=Side(border_style="medium", color="4472C4"))

ws.append(COLUMNS)
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = header_border

ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

# Data rows
for row_idx, p in enumerate(prospects, start=2):
    row_data = [p[col] for col in COLUMNS]
    ws.append(row_data)

    tier = p["Tier"]
    fill = TIER_FILLS[tier]
    font_color = TIER_FONT_COLORS[tier]

    for col_idx, cell in enumerate(ws[row_idx], start=1):
        if col_idx in (1, 2):
            cell.fill = fill
            cell.font = Font(bold=True, color=font_color, size=10)
        else:
            cell.font = Font(size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(
            bottom=Side(border_style="thin", color="CCCCCC"),
            right=Side(border_style="thin", color="CCCCCC"),
        )

    ws.row_dimensions[row_idx].height = 55

# Column widths
col_widths = {
    "A": 10,   # Tier
    "B": 22,   # Segment
    "C": 30,   # Company
    "D": 28,   # Domain
    "E": 15,   # Headcount
    "F": 10,   # Fit Score
    "G": 45,   # Buying Signal
    "H": 25,   # Contact Name
    "I": 25,   # Title
    "J": 40,   # LinkedIn
    "K": 35,   # Email Pattern
    "L": 55,   # Why Now
    "M": 18,   # Email Format Valid
    "N": 16,   # Domain MX Valid
    "O": 22,   # Verified At
}

# Color-code Email Format Valid and Domain MX Valid columns
yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yes_font_color = "276221"
no_font_color = "9C0006"

EMAIL_COL_IDX = COLUMNS.index("Email Format Valid") + 1
MX_VALID_COL_IDX = COLUMNS.index("Domain MX Valid") + 1

for row_idx in range(2, len(prospects) + 2):
    for col_idx in (EMAIL_COL_IDX, MX_VALID_COL_IDX):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value == "yes":
            cell.fill = yes_fill
            cell.font = Font(size=10, color=yes_font_color)
        else:
            cell.fill = no_fill
            cell.font = Font(size=10, color=no_font_color)
        cell.alignment = Alignment(horizontal="center", vertical="top")

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 12

summary_header_fill = PatternFill(start_color="0D1F3C", end_color="0D1F3C", fill_type="solid")

totals = {"Tier 1": 0, "Tier 2": 0, "Tier 3": 0}
seg_totals = {}
for p in prospects:
    totals[p["Tier"]] += 1
    seg_totals[p["Segment"]] = seg_totals.get(p["Segment"], 0) + 1

ws2.append(["ClaimSignal B2B Prospect List — Summary", ""])
ws2["A1"].font = Font(bold=True, size=14, color="2F5496")
ws2["A1"].alignment = Alignment(horizontal="left")
ws2.row_dimensions[1].height = 25
ws2.append([])

ws2.append(["Category", "Count"])
for cell in ws2[3]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = summary_header_fill

ws2.append(["TOTAL PROSPECTS", len(prospects)])
ws2.append([])
for tier, count in totals.items():
    ws2.append([tier, count])
ws2.append([])
for seg, count in seg_totals.items():
    ws2.append([seg, count])
ws2.append([])
ws2.append(["— Email & Domain Verification —", ""])
ws2.append(["Email Format Valid", email_valid_count])
ws2.append(["Email Format Invalid (flagged)", email_invalid_count])
ws2.append(["Domains MX Valid", mx_valid_count])
ws2.append(["Domains MX Invalid (flagged)", mx_invalid_count])
ws2.append(["Verified At (UTC)", verified_at])

wb.save("claimsignal_prospects.xlsx")
print(f"✓ XLSX saved — {len(prospects)} prospects")

# Print breakdown
print(f"\nBreakdown:")
for tier, count in totals.items():
    print(f"  {tier}: {count}")
for seg, count in seg_totals.items():
    print(f"  {seg}: {count}")

# Print top Tier 1 picks
print("\nTop 3 Tier 1 picks:")
tier1 = [p for p in prospects if p["Tier"] == "Tier 1"][:3]
for p in tier1:
    print(f"  • {p['Company']} — {p['Why Now']}")
